from flask import g, request,jsonify
from app.md.models.base import (College_Course,College_Cutoff,College_Data)
from app import db
from flask_restful import Resource
import openpyxl


class College_Data_UploadView(Resource):

    def get(self):
        return {"msg":"college_data"}
    
    def post(self):
        file=request.files['college']
        workbook = openpyxl.load_workbook(file)
        a=[]
        sheet1=[]
        sheet2=[]
        sheet3=[]
        for sheet_name in workbook.sheetnames:
            # sheet = workbook[sheet_name]
            a.append(sheet_name)

        for sheet_name in workbook.sheetnames:
            if a[0]==sheet_name:
                a[0]=workbook[sheet_name]
                for row in a[0].iter_rows(min_row=2, values_only=True):
                    sheet1.append(row)
            if a[1]==sheet_name:
                a[1]=workbook[sheet_name]
                for row in a[1].iter_rows(min_row=2, values_only=True):
                    sheet2.append(row)
            if a[2]==sheet_name:
                a[2]=workbook[sheet_name]
                for row in a[2].iter_rows(min_row=2, values_only=True):
                    sheet3.append(row)
        for college_code1,college_name,city,state,college_type in sheet1:
            data = College_Data(college_code1=college_code1,college_name=college_name,city=city,state=state,college_type=college_type)
            db.session.add(data)
            db.session.flush()
            clg_id=data.id
            for college_code2,course_name,category,no_of_seats in sheet2:
                if college_code1==college_code2:
                    data1 = College_Course(college_code2=college_code2,course_name=course_name,category=category,no_of_seats=no_of_seats,college_id=clg_id)
                    db.session.add(data1)

            for college_code3,course_name2,category2,rank_high,rank_low,marks_high,marks_low,period in sheet3:
                if college_code1==college_code3:
                    data2 = College_Cutoff(college_code3=college_code3,course_name=course_name2,category=category2,rank_high=rank_high,rank_low=rank_low,marks_high=marks_high,marks_low=marks_low,period=period,college_id=clg_id)
                    db.session.add(data2)
        db.session.commit()
        return jsonify({'msg':'data inserted'})
